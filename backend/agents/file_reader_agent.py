# backend/agents/file_reader_agent.py
from backend.app.schemas import AppState, AgentTraceEntry, MeetingLogEntry, File
from datetime import datetime, timezone
import logging
import io
from typing import Any
try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    import docx
except ImportError:
    docx = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    from PIL import Image
    import pytesseract
except ImportError:
    Image = None
    pytesseract = None


# Configure basic logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def log_interaction(state: AppState, decision: str, message: str, level: str = "info"):
    """Helper function to log agent interactions consistently."""
    timestamp = datetime.now(timezone.utc)
    
    state.agent_trace.append(AgentTraceEntry(
        agent="file_reader",
        decision=decision,
        timestamp=timestamp,
    ))
    
    state.meeting_log.append(MeetingLogEntry(
        agent="file_reader",
        message=message,
        timestamp=timestamp
    ))

    if level == "error":
        logger.error(f"File Reader Agent: {message} - Decision: {decision}")
    else:
        logger.info(f"File Reader Agent: {message} - Decision: {decision}")
    state.updated_at = timestamp

def extract_xlsx_content(file_content_bytes: bytes, file_name: str) -> str:
    """
    Extracts text content from XLSX files using openpyxl.
    Returns formatted text representation of the spreadsheet data.
    """
    if not load_workbook:
        return "[Skipped: openpyxl library not available.]"
    
    try:
        workbook = load_workbook(io.BytesIO(file_content_bytes), read_only=True, data_only=True)
        extracted_text = ""
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            extracted_text += f"\\n=== Sheet: {sheet_name} ===\\n"
            
            # Get the used range of cells
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            if max_row == 1 and max_col == 1 and worksheet.cell(1, 1).value is None:
                extracted_text += "[Empty sheet]\\n"
                continue
            
            # Extract data row by row
            for row_num in range(1, min(max_row + 1, 1000)):  # Limit to 1000 rows to avoid huge outputs
                row_data = []
                has_data = False
                
                for col_num in range(1, min(max_col + 1, 50)):  # Limit to 50 columns
                    cell_value = worksheet.cell(row_num, col_num).value
                    if cell_value is not None:
                        has_data = True
                        # Convert cell value to string, handling different data types
                        if isinstance(cell_value, (int, float)):
                            cell_str = str(cell_value)
                        else:
                            # Handle datetime objects
                            from datetime import date, datetime as dt
                            if isinstance(cell_value, (date, dt)):
                                try:
                                    cell_str = cell_value.strftime('%Y-%m-%d %H:%M:%S')
                                except AttributeError:
                                    cell_str = str(cell_value)
                            else:
                                cell_str = str(cell_value)
                        row_data.append(cell_str)
                    else:
                        row_data.append("")
                
                # Only add row if it has data
                if has_data:
                    # Remove trailing empty cells
                    while row_data and row_data[-1] == "":
                        row_data.pop()
                    
                    if row_data:  # If there's still data after cleanup
                        extracted_text += " | ".join(row_data) + "\\n"
            
            extracted_text += "\\n"
        
        workbook.close()
        return extracted_text
        
    except Exception as e:
        return f"[Error extracting XLSX content: {str(e)}]"

def handle(state_dict: dict) -> dict:
    state = AppState(**state_dict)
    log_interaction(state, "Starting file processing", "File Reader Agent invoked.")

    if state.processed_files_content is None:
        state.processed_files_content = {}

    if state.files:
        for file_data_model in state.files:
            file_name = file_data_model.name
            file_content_bytes = file_data_model.content
            content_type = file_data_model.metadata.get('content_type', 'application/octet-stream')
            
            parsed_text_content = f"Content of {file_name} (type: {content_type}):\\n"

            try:
                if file_content_bytes is None:
                    parsed_text_content += "[Error: File content is missing.]"
                    log_interaction(state, f"Missing content for {file_name}", f"File {file_name} has no content.", level="error")
                    state.processed_files_content[file_name] = parsed_text_content
                    continue

                if content_type == 'text/plain':
                    try:
                        parsed_text_content += file_content_bytes.decode('utf-8')
                        log_interaction(state, f"Processed text file {file_name}", f"Successfully decoded {file_name} as UTF-8 text.")
                    except UnicodeDecodeError:
                        parsed_text_content += "[Error: Could not decode content as UTF-8.]"
                        log_interaction(state, f"Encoding error for {file_name}", f"Could not decode {file_name} as UTF-8.", level="error")
                elif content_type == 'application/pdf':
                    if PyPDF2:
                        try:
                            pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content_bytes))
                            text = ""
                            for page_num in range(len(pdf_reader.pages)):
                                text += pdf_reader.pages[page_num].extract_text()
                            parsed_text_content += text
                            log_interaction(state, f"Processed PDF file {file_name}", f"Successfully extracted text from PDF {file_name}.")
                        except Exception as e:
                            parsed_text_content += f"[Error extracting PDF content: {str(e)}]"
                            log_interaction(state, f"PDF extraction error for {file_name}", f"Error extracting text from PDF {file_name}: {str(e)}", level="error")
                    else:
                        parsed_text_content += "[Skipped: PyPDF2 library not available.]"
                        log_interaction(state, f"Skipped PDF file {file_name}", "PyPDF2 library not available.", level="warning")
                elif content_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
                    if docx:
                        try:
                            document = docx.Document(io.BytesIO(file_content_bytes))
                            text = "\\n".join([para.text for para in document.paragraphs])
                            parsed_text_content += text
                            log_interaction(state, f"Processed DOCX file {file_name}", f"Successfully extracted text from DOCX {file_name}.")
                        except Exception as e:
                            parsed_text_content += f"[Error extracting DOCX content: {str(e)}]"
                            log_interaction(state, f"DOCX extraction error for {file_name}", f"Error extracting text from DOCX {file_name}: {str(e)}", level="error")
                    else:
                        parsed_text_content += "[Skipped: python-docx library not available.]"
                        log_interaction(state, f"Skipped DOCX file {file_name}", "python-docx library not available.", level="warning")
                elif content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                    # Handle XLSX files
                    if load_workbook:
                        try:
                            xlsx_content = extract_xlsx_content(file_content_bytes, file_name)
                            parsed_text_content += xlsx_content
                            log_interaction(state, f"Processed XLSX file {file_name}", f"Successfully extracted data from XLSX {file_name}.")
                        except Exception as e:
                            parsed_text_content += f"[Error extracting XLSX content: {str(e)}]"
                            log_interaction(state, f"XLSX extraction error for {file_name}", f"Error extracting data from XLSX {file_name}: {str(e)}", level="error")
                    else:
                        parsed_text_content += "[Skipped: openpyxl library not available.]"
                        log_interaction(state, f"Skipped XLSX file {file_name}", "openpyxl library not available.", level="warning")
                elif content_type.startswith('image/'):
                    if Image and pytesseract:
                        try:
                            image = Image.open(io.BytesIO(file_content_bytes))
                            text = pytesseract.image_to_string(image)
                            parsed_text_content += text
                            log_interaction(state, f"Processed image file {file_name} with OCR", f"Successfully extracted text from image {file_name} using Tesseract OCR.")
                        except Exception as e:
                            parsed_text_content += f"[Error processing image with OCR: {str(e)}]"
                            log_interaction(state, f"Image OCR error for {file_name}", f"Error processing image {file_name} with OCR: {str(e)}", level="error")
                    else:
                        parsed_text_content += "[Skipped: Pillow or pytesseract library not available.]"
                        log_interaction(state, f"Skipped image file {file_name}", "Pillow or pytesseract library not available.", level="warning")
                else:
                    parsed_text_content += "[Unsupported file type for content extraction.]"
                    log_interaction(state, f"Unsupported file type {content_type} for {file_name}", f"File {file_name} has an unsupported content type: {content_type}.", level="warning")
                
                state.processed_files_content[file_name] = parsed_text_content

            except Exception as e:
                error_msg = f"Error processing file {file_name}: {str(e)}"
                log_interaction(state, f"Error processing file {file_name}", error_msg, level="error")
                state.processed_files_content[file_name] = f"[Error processing file: {str(e)}]"
    else:
        log_interaction(state, "No files to process", "No files found in state.files.")

    log_interaction(state, "File processing complete", "File Reader Agent finished.")
    return state.model_dump()
